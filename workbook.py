import openpyxl

#%%
wb = openpyxl.load_workbook('./hello.xlsx')
print('现有工作表名称为', wb.sheetnames)
ws1 = wb.worksheets[0]
print(ws1.cell(3, 2).value)

#%%
# def judge_sheet(a_workbook,a_sheet_num):
#     if a_sheet_num  > a_workbook.worksheets
#         ws2 = wb.create_sheet('测试表2')
#         print("创建测试表2成功")
#         print(wb.sheetnames)
#     else :
#         print('表格现在存在的数量大于预期数量')
#
# judge_sheet(wb, sheet_number)
# wb.save('hello.xlsx')
# # %%
